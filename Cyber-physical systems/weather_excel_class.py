import calendar

from interpolate_methods import interpol_methods
import json
import os
import re

from openpyxl import load_workbook
from openpyxl.comments import Comment


class WeatherExcel:
    """Class describes working with Excel files in weather format.

    Columns: Day, Time, Temperature, Direction, Speed,
    WeatherCode, Clouds, Visibility, Pressure, CloudBase."""

    def __init__(self, year: int, month: int, input_path: str, save_path: str,
                 table_head_path: str, comments_path: str, directions_path: str,
                 direction_numbers_path: str, months_path: str):
        if not isinstance(year, int):
            raise TypeError("year must be an integer")
        if year <= 0:
            raise ValueError("year must be a positive")
        if not isinstance(month, int):
            raise TypeError("month must be an integer")
        if not 1 <= month <= 12:
            raise ValueError("month must be in range from 1 to 12")

        self.__year = year
        self.__month = month

        if self.month in (1, 3, 5, 7, 8, 10, 12):
            self.__rows = 31 * 48 + 1
        elif self.month in (4, 6, 9, 11):
            self.__rows = 30 * 48 + 1
        elif calendar.isleap(self.year):
            self.__rows = 29 * 48 + 1
        else:
            self.__rows = 28 * 48 + 1

        self.__output_path = os.path.join(
            save_path, str(self.year) + "-" + str(self.month) + ".xlsx")

        # make a copy of input file
        load_workbook(filename=input_path).save(self.output_path)

        self.__wb = load_workbook(filename=self.output_path)
        self.__ws = self.wb.active

        with open(table_head_path) as table_head_file:
            self.__table_head = json.load(table_head_file)
        with open(comments_path) as comments_file:
            self.__comments = json.load(comments_file)
        with open(directions_path) as directions_file:
            self.__directions = json.load(directions_file)
        with open(direction_numbers_path) as direction_numbers_file:
            self.__direction_numbers = json.load(direction_numbers_file)
        with open(months_path) as months_file:
            self.__months = json.load(months_file)

    @property
    def year(self):
        return self.__year

    @property
    def month(self):
        return self.__month

    @property
    def rows(self):
        return self.__rows

    @property
    def output_path(self):
        return self.__output_path

    @property
    def wb(self):
        return self.__wb

    @property
    def ws(self):
        return self.__ws

    @property
    def table_head(self):
        return self.__table_head

    @property
    def comments(self):
        return self.__comments

    @property
    def directions(self):
        return self.__directions

    @property
    def direction_numbers(self):
        return self.__direction_numbers

    @property
    def months(self):
        return self.__months

    def update_head(self, lang: str):
        if not isinstance(lang, str):
            raise TypeError("lang must be a string")
        if lang not in self.table_head.keys():
            raise ValueError("the lang not supported")

        self.ws["A1"] = self.table_head[lang]["Day"]
        self.ws["B1"] = self.table_head[lang]["Time"]
        self.ws["C1"] = self.table_head[lang]["Temperature"]
        self.ws["D1"] = self.table_head[lang]["Direction"]
        self.ws["E1"] = self.table_head[lang]["Speed"]
        self.ws["F1"] = self.table_head[lang]["Weather code"]
        self.ws["G1"] = self.table_head[lang]["Clouds"]
        self.ws["H1"] = self.table_head[lang]["Visibility"]
        self.ws["I1"] = self.table_head[lang]["Pressure"]
        self.ws["J1"] = self.table_head[lang]["Cloud base"]

        self.wb.save(self.output_path)

    def update_comments(self, lang):
        if not isinstance(lang, str):
            raise TypeError("lang must be a string")
        if lang not in self.table_head.keys():
            raise ValueError("the lang not supported")

        self.ws["A1"].comment = Comment(self.comments[lang]["Day"], "Author")
        self.ws["B1"].comment = Comment(self.comments[lang]["Time"], "Author")
        self.ws["C1"].comment = Comment(self.comments[lang]["Temperature"], "Author")
        self.ws["E1"].comment = Comment(self.comments[lang]["Direction"], "Author")
        self.ws["F1"].comment = Comment(self.comments[lang]["Weather code"], "Author")
        self.ws["G1"].comment = Comment(self.comments[lang]["Clouds"], "Author")
        self.ws["H1"].comment = Comment(self.comments[lang]["Visibility"], "Author")
        self.ws["I1"].comment = Comment(self.comments[lang]["Pressure"], "Author")
        self.ws["J1"].comment = Comment(self.comments[lang]["Cloud base"], "Author")

        self.wb.save(self.output_path)

    def update_temperature(self, interpolate_method: str):
        if not isinstance(interpolate_method, str):
            raise TypeError("interpolate method must be a string")
        if interpolate_method not in interpol_methods.keys():
            raise ValueError("method must be one of the interpol methods")
        values = []
        values_indexes = []
        null_indexes = []

        for col in self.ws.iter_cols(min_row=2, max_row=self.rows,
                                     min_col=3, max_col=3, values_only=True):
            for i, val in enumerate(col):
                if val is not None:
                    values.append(val)
                    values_indexes.append(i)
                else:
                    null_indexes.append(i)

        if null_indexes:
            interpolated = tuple(map(
                round, interpol_methods[interpolate_method](
                    null_indexes, values_indexes, values)))
            for i, val in enumerate(interpolated):
                self.ws["C" + str(null_indexes[i] + 2)] = val

        self.wb.save(self.output_path)

    def update_speed(self, interpolate_method: str):
        if not isinstance(interpolate_method, str):
            raise TypeError("interpolate method must be a string")
        if interpolate_method not in interpol_methods.keys():
            raise ValueError("method must be one of the interpol methods")
        values = []
        values_indexes = []
        null_indexes = []

        for col in self.ws.iter_cols(min_row=2, max_row=self.rows,
                                     min_col=5, max_col=5, values_only=True):
            for i, val in enumerate(col):
                if val is not None:
                    values.append(val)
                    values_indexes.append(i)
                else:
                    null_indexes.append(i)

        if null_indexes:
            interpolated = tuple(map(
                round, interpol_methods[interpolate_method](
                    null_indexes, values_indexes, values)))
            for i, val in enumerate(interpolated):
                self.ws["E" + str(null_indexes[i] + 2)] = val

        self.wb.save(self.output_path)

    def update_direction(self, interpolate_method: str):
        if not isinstance(interpolate_method, str):
            raise TypeError("interpolate method must be a string")
        if interpolate_method not in interpol_methods.keys():
            raise ValueError("method must be one of the interpol methods")
        values = []
        values_indexes = []
        null_indexes = []

        for col in self.ws.iter_cols(min_row=2, max_row=self.rows,
                                     min_col=4, max_col=4, values_only=True):
            for i, val in enumerate(col):
                if val == "Переменный":
                    self.ws["D" + str(i + 2)] = "Змінний"
                elif val is not None:
                    values.append(self.direction_numbers[val])
                    values_indexes.append(i)
                    self.ws["D" + str(i + 2)] = self.directions[val]
                elif not self.ws["E" + str(i + 2)].value:
                    self.ws["D" + str(i + 2)] = "Штиль"
                else:
                    null_indexes.append(i)

        if null_indexes:
            interpolated = tuple(map(
                round, interpol_methods[interpolate_method](
                    null_indexes, values_indexes, values)))
            for i, val in enumerate(interpolated):
                self.ws["D" + str(null_indexes[i] + 2)] = self.directions[val]

        self.wb.save(self.output_path)

    def update_weather_code(self):
        for col in self.ws.iter_cols(min_row=2, max_row=self.rows,
                                     min_col=6, max_col=6, values_only=True):
            for i, val in enumerate(col):
                if not val:
                    self.ws["F" + str(i + 2)] = "CL"

        self.wb.save(self.output_path)

    def update_clouds(self, interpolate_method: str):
        if not isinstance(interpolate_method, str):
            raise TypeError("interpolate method must be a string")
        if interpolate_method not in interpol_methods.keys():
            raise ValueError("method must be one of the interpol methods")
        values = []
        values_indexes = []
        null_indexes = []

        for col in self.ws.iter_cols(min_row=2, max_row=self.rows,
                                     min_col=7, max_col=7, values_only=True):
            for i, val in enumerate(col):
                if val is not None:
                    values.append(val)
                    values_indexes.append(i)
                else:
                    null_indexes.append(i)

        if null_indexes:
            interpolated = tuple(map(
                round, interpol_methods[interpolate_method](
                    null_indexes, values_indexes, values)))
            for i, val in enumerate(interpolated):
                self.ws["G" + str(null_indexes[i] + 2)] = val

        self.wb.save(self.output_path)

    def update_visibility(self, interpolate_method: str):
        if not isinstance(interpolate_method, str):
            raise TypeError("interpolate method must be a string")
        if interpolate_method not in interpol_methods.keys():
            raise ValueError("method must be one of the interpol methods")
        values = []
        values_indexes = []
        null_indexes = []

        for col in self.ws.iter_cols(min_row=2, max_row=self.rows,
                                     min_col=8, max_col=8, values_only=True):
            for i, val in enumerate(col):
                if val is not None:
                    whole_part = re.search(r"\d+.\d+", val)
                    if whole_part:
                        whole_part = float(whole_part.group())
                    else:
                        whole_part = int(re.search(r"\d+", val).group())
                        float_part = re.search(r"[^\d,.-]+", val).group()
                        whole_part += self.months[float_part.lower()] / 10
                    values.append(whole_part)
                    values_indexes.append(i)
                    self.ws["H" + str(i + 2)] = whole_part
                else:
                    null_indexes.append(i)

        if null_indexes:
            interpolated = list(round(num, 1) for num in
                                interpol_methods[interpolate_method](
                null_indexes, values_indexes, values))
            for i, val in enumerate(interpolated):
                self.ws["H" + str(null_indexes[i] + 2)] = val

        self.wb.save(self.output_path)

    def update_pressure(self, interpolate_method: str):
        if not isinstance(interpolate_method, str):
            raise TypeError("interpolate method must be a string")
        if interpolate_method not in interpol_methods.keys():
            raise ValueError("method must be one of the interpol methods")
        values = []
        values_indexes = []
        null_indexes = []

        for col in self.ws.iter_cols(min_row=2, max_row=self.rows,
                                     min_col=10, max_col=10, values_only=True):
            for i, val in enumerate(col):
                if val is not None:
                    values.append(val)
                    values_indexes.append(i)
                else:
                    null_indexes.append(i)

        for i, val in enumerate(values_indexes):
            self.ws["I" + str(val + 2)] = values[i]

        if null_indexes:
            interpolated = tuple(map(
                round, interpol_methods[interpolate_method](
                    null_indexes, values_indexes, values)))
            for i, val in enumerate(interpolated):
                self.ws["I" + str(null_indexes[i] + 2)] = val

        self.wb.save(self.output_path)

    def update_cloud_base(self, interpolate_method: str):
        if not isinstance(interpolate_method, str):
            raise TypeError("interpolate method must be a string")
        if interpolate_method not in interpol_methods.keys():
            raise ValueError("method must be one of the interpol methods")
        values = []
        values_indexes = []
        null_indexes = []

        for col in self.ws.iter_cols(min_row=2, max_row=self.rows,
                                     min_col=11, max_col=11, values_only=True):
            for i, val in enumerate(col):
                if val is not None:
                    values.append(val)
                    values_indexes.append(i)
                else:
                    null_indexes.append(i)

        for i, val in enumerate(values_indexes):
            self.ws["J" + str(val + 2)] = values[i]

        if null_indexes:
            interpolated = tuple(map(
                round, interpol_methods[interpolate_method](
                    null_indexes, values_indexes, values)))
            for i, val in enumerate(interpolated):
                self.ws["J" + str(null_indexes[i] + 2)] = val

        self.ws.delete_cols(11, 1)
        self.wb.save(self.output_path)
